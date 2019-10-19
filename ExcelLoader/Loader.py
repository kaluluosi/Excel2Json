import os

import openpyxl
from openpyxl.cell import Cell
from openpyxl.styles.colors import Color
from openpyxl.workbook import Workbook
from openpyxl.worksheet import worksheet


class RawData(object):
    """
    原始数据
    """

    def __init__(self, cfg_name: str, type_names: list, field_names: list, data: dict):
        self.type_names = type_names
        self.field_names = field_names
        self.data = data
        self.cfg_file_name = cfg_name


class Setting(object):
    """
    标志设置
    每个配置表左上角是标志信息，标志了类型行、字段行、结束行、配置名这几个属性
    """

    def __init__(self, cfg_name, type_row: int, field_row: int, data_range: tuple):
        self.__cfg_name = cfg_name
        self.__type_row = type_row
        self.__field_row = field_row
        self.__range = data_range

    @property
    def type_row(self):
        return self.__type_row

    @property
    def field_row(self):
        return self.__field_row

    @property
    def range(self):
        return self.__range

    @property
    def cfg_name(self):
        return self.__cfg_name


class ExcelLoader(object):

    def __init__(self, file_name: str):
        self.file_name = file_name
        self.workbook = self.open_workbook(self.file_name)

    def open_workbook(self) -> Workbook:
        raise NotImplementedError()

    def get_raw_data(self, sheet_index) -> RawData:
        """获取某个sheet的原始数据"""
        return

    @property
    def all_raw_data(self):
        for i, sheet in enumerate(self.workbook.worksheets):
            yield self.get_raw_data(i)

    def setting(self, sheet_index) -> Setting:
        raise NotImplementedError()

    def close(self):
        self.workbook.close()


def Parse(t: str, value):
    if t == 'int':
        try:
            return int(value)
        except:
            return -1

    if t == 'string':
        return str(value)

    if t.startswith('dict'):
        try:
            return eval(value)
        except:
            return None

    return eval(value)


class XLSXLoader(ExcelLoader):

    def __init__(self, file_name: str):
        super(XLSXLoader, self).__init__(file_name)

    def open_workbook(self, file_name: str) -> worksheet:
        workbook = openpyxl.open(file_name)  # type: Workbook
        return workbook

    def setting(self, sheet_index=0) -> Setting:
        """设置数据"""
        sheet = self.workbook.worksheets[sheet_index]  # type: worksheet

        # 写死读取设置
        type_color = sheet.cell(1, 2).fill.fgColor  # type: Color
        field_name_color = sheet.cell(2, 2).fill.fgColor  # type:Color
        end_color = sheet.cell(3, 2).fill.fgColor  # type: Color
        cfg_name = sheet.cell(4, 2).value

        # row最小是1，而range是[ )左闭右开区间，所以右边要+1
        for row in range(1, sheet.max_row + 1):

            cell = sheet.cell(row, 1)  # type:Cell

            if cell.fill.fgColor == type_color:
                type_row = row
            if cell.fill.fgColor == field_name_color:
                field_row = row
            if cell.fill.fgColor == end_color:
                end_row = row
                break

        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(type_row, col)

            if cell.fill.fgColor != type_color:
                data_max_col = col
                break

        return Setting(cfg_name, type_row, field_row, (end_row, data_max_col))

    def get_raw_data(self, sheet_index=0):
        sheet = self.workbook.worksheets[sheet_index]  # type: worksheet
        setting = self.setting(sheet_index)
        cfg_name = setting.cfg_name

        # 获取字段名
        # 获取字段类型名，默认int
        field_names = ['']
        type_names = ['']

        for col in range(1, setting.range[1]):
            field_cell = sheet.cell(setting.field_row, col)  # type: Cell
            type_cell = sheet.cell(setting.type_row, col)  # type: Cell
            field_names.append(field_cell.value)
            type_names.append(type_cell.value or 'int')  # 如果类型为空默认当int

        # 遍历每行数据，用字典的方式组织
        data = {}
        for raw in range(setting.field_row + 1, setting.range[0]):
            item = {}
            for col in range(1, setting.range[1]):
                cell = sheet.cell(raw, col)
                field_name = field_names[col]
                # 跳过空id的数据
                if field_name == 'id' and cell.value == '':
                    continue
                item[field_name] = Parse(type_names[col], cell.value)

            data[item['id']] = item

        raw_data = RawData(cfg_name, type_names, field_names, data)

        return raw_data


class XLSLoader(ExcelLoader):

    def __init__(self, filename: str):
        super(XLSLoader, self).__init__(filename)


def create_loader(filename: str) -> ExcelLoader:
    extension = os.path.splitext(filename)[1]

    if extension == '.xlsx':
        loader = XLSXLoader(filename)
        return loader
    elif extension == '.xls':
        loader = XLSLoader(filename)
        return loader
